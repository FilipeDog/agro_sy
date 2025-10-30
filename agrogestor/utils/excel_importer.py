# -*- coding: utf-8 -*-
"""
Importador de Dados via Excel
"""
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from datetime import datetime
try:
    import ttkbootstrap as ttk
    from ttkbootstrap.constants import *
except ImportError:
    from tkinter.constants import *
    import tkinter.ttk as ttk


class ExcelImporterWindow:
    """Janela para importação de dados via Excel"""

    def __init__(self, parent, db_manager):
        self.parent = parent
        self.db_manager = db_manager
        self.workbook = None
        self.current_sheet = None
        self.column_mapping = {}
        
        self.create_widgets()

    def create_widgets(self):
        main_container = ttk.Frame(self.parent)
        main_container.pack(fill=BOTH, expand=YES)

        ttk.Label(main_container, text="Importar Dados do Excel", 
                 font=("Helvetica", 16, "bold")).pack(padx=20, pady=10)

        # Frame de seleção de arquivo
        file_frame = ttk.LabelFrame(main_container, text="1. Selecionar Arquivo", padding=20)
        file_frame.pack(fill=X, padx=20, pady=10)

        self.file_label = ttk.Label(file_frame, text="Nenhum arquivo selecionado")
        self.file_label.pack(side=LEFT, padx=5)

        ttk.Button(file_frame, text="Selecionar Excel", 
                  command=self.select_file).pack(side=RIGHT, padx=5)

        # Frame de seleção de planilha e tipo
        config_frame = ttk.LabelFrame(main_container, text="2. Configurar Importação", padding=20)
        config_frame.pack(fill=X, padx=20, pady=10)

        ttk.Label(config_frame, text="Tipo de Dados:").grid(row=0, column=0, sticky=W, pady=5)
        self.tipo_combo = ttk.Combobox(config_frame, width=25, state="readonly", values=[
            "Clientes",
            "Fornecedores",
            "Funcionários",
            "Animais",
            "Talhões",
            "Despesas",
            "Receitas",
            "Tratos Culturais",
            "Colheitas"
        ])
        self.tipo_combo.grid(row=0, column=1, sticky=W, pady=5, padx=5)
        self.tipo_combo.bind('<<ComboboxSelected>>', self.update_mapping_template)

        ttk.Label(config_frame, text="Planilha:").grid(row=1, column=0, sticky=W, pady=5)
        self.sheet_combo = ttk.Combobox(config_frame, width=25, state="readonly")
        self.sheet_combo.grid(row=1, column=1, sticky=W, pady=5, padx=5)
        self.sheet_combo.bind('<<ComboboxSelected>>', self.load_sheet_preview)

        ttk.Label(config_frame, text="Linha inicial (dados):").grid(row=2, column=0, sticky=W, pady=5)
        self.start_row_entry = ttk.Entry(config_frame, width=10)
        self.start_row_entry.grid(row=2, column=1, sticky=W, pady=5, padx=5)
        self.start_row_entry.insert(0, "2")

        # Frame de preview
        preview_frame = ttk.LabelFrame(main_container, text="3. Preview dos Dados", padding=20)
        preview_frame.pack(fill=BOTH, expand=YES, padx=20, pady=10)

        # Treeview para preview
        tree_frame = ttk.Frame(preview_frame)
        tree_frame.pack(fill=BOTH, expand=YES)

        vsb = ttk.Scrollbar(tree_frame, orient=VERTICAL)
        hsb = ttk.Scrollbar(tree_frame, orient=HORIZONTAL)

        self.preview_tree = ttk.Treeview(tree_frame, 
                                        yscrollcommand=vsb.set,
                                        xscrollcommand=hsb.set,
                                        height=10)
        
        vsb.config(command=self.preview_tree.yview)
        hsb.config(command=self.preview_tree.xview)

        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Botões
        btn_frame = ttk.Frame(main_container)
        btn_frame.pack(fill=X, padx=20, pady=20)

        ttk.Button(btn_frame, text="Importar Dados", 
                  command=self.import_data, width=20).pack(side=RIGHT, padx=5)
        ttk.Button(btn_frame, text="Baixar Modelo Excel", 
                  command=self.download_template, width=20).pack(side=RIGHT, padx=5)

    def select_file(self):
        """Seleciona arquivo Excel"""
        filename = filedialog.askopenfilename(
            title="Selecionar arquivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )

        if filename:
            try:
                self.workbook = openpyxl.load_workbook(filename, data_only=True)
                self.file_label.config(text=filename.split('/')[-1])
                
                # Carregar nomes das planilhas
                self.sheet_combo['values'] = self.workbook.sheetnames
                if self.workbook.sheetnames:
                    self.sheet_combo.set(self.workbook.sheetnames[0])
                    self.load_sheet_preview()
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao abrir arquivo: {str(e)}")

    def load_sheet_preview(self, event=None):
        """Carrega preview da planilha"""
        if not self.workbook or not self.sheet_combo.get():
            return

        try:
            self.current_sheet = self.workbook[self.sheet_combo.get()]
            
            # Limpar preview anterior
            self.preview_tree.delete(*self.preview_tree.get_children())
            
            # Configurar colunas
            first_row = list(self.current_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            columns = [str(col) if col else f"Col{i+1}" for i, col in enumerate(first_row)]
            
            self.preview_tree['columns'] = columns
            self.preview_tree['show'] = 'headings'
            
            for col in columns:
                self.preview_tree.heading(col, text=col)
                self.preview_tree.column(col, width=150)
            
            # Carregar primeiras 20 linhas
            for idx, row in enumerate(self.current_sheet.iter_rows(min_row=2, max_row=21, values_only=True)):
                if any(row):  # Se há algum valor na linha
                    self.preview_tree.insert('', END, values=row)
                    
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar preview: {str(e)}")

    def update_mapping_template(self, event=None):
        """Atualiza template de mapeamento baseado no tipo"""
        tipo = self.tipo_combo.get()

        # Templates com campos obrigatórios e opcionais
        # Formato: (nome_campo, obrigatório, descrição_exemplo)
        templates = {
            "Clientes": [
                ("nome", True, "João Silva"),
                ("cpf_cnpj", False, "123.456.789-00"),
                ("email", False, "joao@email.com"),
                ("telefone", False, "(11) 99999-9999"),
                ("cidade", False, "São Paulo"),
                ("uf", False, "SP"),
                ("endereco", False, "Rua das Flores, 123"),
                ("cep", False, "01234-567"),
                ("observacoes", False, "Cliente VIP")
            ],
            "Fornecedores": [
                ("nome", True, "Fornecedor ABC"),
                ("cpf_cnpj", False, "12.345.678/0001-90"),
                ("email", False, "contato@fornecedor.com"),
                ("telefone", False, "(11) 3333-3333"),
                ("cidade", False, "São Paulo"),
                ("uf", False, "SP"),
                ("endereco", False, "Av. Comercial, 456"),
                ("cep", False, "01234-567"),
                ("observacoes", False, "")
            ],
            "Funcionários": [
                ("nome", True, "Maria Santos"),
                ("cpf", True, "123.456.789-00"),
                ("rg", False, "12.345.678-9"),
                ("telefone", False, "(11) 98888-8888"),
                ("email", False, "maria@email.com"),
                ("cargo", False, "Operador"),
                ("setor", False, "Campo"),
                ("salario", False, "2500.00"),
                ("data_admissao", False, "01/01/2024"),
                ("data_nascimento", False, "15/05/1990")
            ],
            "Animais": [
                ("brinco", True, "001"),
                ("lote", False, "Lote A"),
                ("sexo", False, "Macho"),
                ("tipo_animal", False, "Boi"),
                ("raca", False, "Nelore"),
                ("data_nascimento", False, "15/03/2022"),
                ("peso_atual", False, "450.5"),
                ("pasto", False, "Pasto Norte"),
                ("status", False, "Ativo"),
                ("origem", False, "Compra"),
                ("nome", False, "Mimoso")
            ],
            "Talhões": [
                ("codigo", True, "T001"),
                ("nome", True, "Talhão Norte"),
                ("area_hectares", True, "5.5"),
                ("localizacao", False, "Fazenda Santa Clara"),
                ("variedade", False, "Banana Nanica"),
                ("data_plantio", False, "10/01/2023"),
                ("espacamento", False, "3m x 2m"),
                ("densidade_plantas_ha", False, "1667"),
                ("situacao", False, "Ativo")
            ],
            "Despesas": [
                ("data_gasto", True, "20/01/2024"),
                ("descricao", True, "Compra de ração"),
                ("tipo_despesa", True, "Ração"),
                ("fornecedor", False, "Agropecuária Silva"),
                ("quantidade", False, "50"),
                ("valor_unitario", False, "80.00"),
                ("valor", False, "4000.00"),
                ("desconto", False, "0"),
                ("forma_pagamento", False, "PIX"),
                ("pago", False, "1"),
                ("numero_nota", False, "NF-12345")
            ],
            "Receitas": [
                ("data_venda", True, "20/01/2024"),
                ("descricao", True, "Venda de gado"),
                ("tipo_receita", False, "Venda de Animal"),
                ("cliente", False, "João Comprador"),
                ("quantidade", False, "5"),
                ("valor_unitario", False, "2500.00"),
                ("valor_total", False, "12500.00"),
                ("desconto", False, "0"),
                ("forma_pagamento", False, "Transferência"),
                ("pago", False, "1"),
                ("numero_nota", False, "NF-54321")
            ],
            "Tratos Culturais": [
                ("talhao", True, "Talhão Norte"),
                ("tipo_trato", True, "Adubação"),
                ("data_execucao", True, "15/02/2024"),
                ("produto_utilizado", False, "NPK 20-05-20"),
                ("quantidade", False, "100"),
                ("unidade", False, "kg"),
                ("custo", False, "350.00"),
                ("responsavel", False, "João Silva"),
                ("proxima_aplicacao", False, "15/04/2024")
            ],
            "Colheitas": [
                ("talhao", True, "Talhão Norte"),
                ("data_colheita", True, "20/03/2024"),
                ("quantidade_kg", True, "1500"),
                ("quantidade_caixas", False, "75"),
                ("peso_medio_cacho", False, "20"),
                ("classificacao_a_kg", False, "900"),
                ("classificacao_b_kg", False, "450"),
                ("classificacao_c_kg", False, "150"),
                ("custo_colheita", False, "300.00"),
                ("responsavel", False, "Maria Santos"),
                ("destino", False, "Venda")
            ]
        }

        self.current_template = templates.get(tipo, [])

    def import_data(self):
        """Importa os dados"""
        if not self.current_sheet or not self.tipo_combo.get():
            messagebox.showwarning("Aviso", "Selecione um arquivo e tipo de dados!")
            return

        try:
            start_row = int(self.start_row_entry.get())
        except:
            messagebox.showerror("Erro", "Linha inicial inválida!")
            return

        tipo = self.tipo_combo.get()
        
        # Confirmar
        if not messagebox.askyesno("Confirmar", 
                                   f"Importar dados como {tipo}?\n\n"
                                   f"Esta ação irá adicionar os registros ao banco de dados."):
            return

        try:
            # Pegar cabeçalhos
            headers = [cell.value for cell in self.current_sheet[1]]

            imported = 0
            errors = []
            error_details = []

            for idx, row in enumerate(self.current_sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if not any(row):
                    continue

                try:
                    data = dict(zip(headers, row))
                    self.import_row(tipo, data)
                    imported += 1
                except ValueError as e:
                    # Erro detalhado do db_manager
                    error_msg = str(e)
                    if "no column named" in error_msg.lower():
                        # Extrair nome da coluna problemática
                        errors.append(f"Linha {idx}: Coluna inexistente no banco de dados")
                        error_details.append(f"Linha {idx}: {error_msg[:200]}")
                    elif "not null" in error_msg.lower():
                        errors.append(f"Linha {idx}: Campo obrigatório faltando")
                        error_details.append(f"Linha {idx}: {error_msg[:200]}")
                    else:
                        errors.append(f"Linha {idx}: {error_msg[:100]}")
                        error_details.append(f"Linha {idx}: {error_msg[:200]}")
                except Exception as e:
                    errors.append(f"Linha {idx}: {str(e)[:100]}")
                    error_details.append(f"Linha {idx}: {str(e)[:200]}")

            # Mostrar resultado
            if imported > 0:
                msg = f"✓ Importação concluída!\n\n{imported} registros importados com sucesso"
                if errors:
                    msg += f"\n⚠ {len(errors)} erros encontrados"
                messagebox.showinfo("Resultado da Importação", msg)
            else:
                messagebox.showwarning("Aviso", "Nenhum registro foi importado!")

            # Mostrar detalhes dos erros
            if errors:
                error_window = tk.Toplevel(self.parent)
                error_window.title("Detalhes dos Erros")
                error_window.geometry("700x500")

                ttk.Label(error_window, text=f"Erros encontrados: {len(errors)}",
                         font=("Helvetica", 12, "bold")).pack(padx=20, pady=10)

                text_frame = ttk.Frame(error_window)
                text_frame.pack(fill=BOTH, expand=YES, padx=20, pady=10)

                text_widget = tk.Text(text_frame, wrap=tk.WORD, height=20)
                scrollbar = ttk.Scrollbar(text_frame, command=text_widget.yview)
                text_widget.configure(yscrollcommand=scrollbar.set)

                text_widget.pack(side=LEFT, fill=BOTH, expand=YES)
                scrollbar.pack(side=RIGHT, fill=Y)

                # Adicionar erros ao widget de texto
                for error_detail in error_details[:50]:  # Limitar a 50 erros
                    text_widget.insert(tk.END, f"{error_detail}\n\n")

                text_widget.config(state=tk.DISABLED)

                ttk.Button(error_window, text="Fechar",
                          command=error_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro crítico na importação:\n\n{str(e)}")

    def import_row(self, tipo, data):
        """Importa uma linha de dados com mapeamento inteligente de foreign keys"""
        # Limpar dados vazios
        clean_data = {k: v for k, v in data.items() if v is not None and str(v).strip() != ''}

        if tipo == "Clientes":
            self.db_manager.insert('clientes', clean_data)

        elif tipo == "Fornecedores":
            self.db_manager.insert('fornecedores', clean_data)

        elif tipo == "Funcionários":
            # Converter datas
            if 'data_nascimento' in clean_data:
                clean_data['data_nascimento'] = self.parse_date(clean_data['data_nascimento'])
            if 'data_admissao' in clean_data:
                clean_data['data_admissao'] = self.parse_date(clean_data['data_admissao'])
            self.db_manager.insert('funcionarios', clean_data)

        elif tipo == "Animais":
            # Converter datas
            if 'data_nascimento' in clean_data:
                clean_data['data_nascimento'] = self.parse_date(clean_data['data_nascimento'])
            if 'data_entrada' in clean_data:
                clean_data['data_entrada'] = self.parse_date(clean_data['data_entrada'])
            if 'data_compra' in clean_data:
                clean_data['data_compra'] = self.parse_date(clean_data['data_compra'])
            if 'data_desmama' in clean_data:
                clean_data['data_desmama'] = self.parse_date(clean_data['data_desmama'])

            # Mapear foreign keys
            clean_data = self.map_foreign_keys(clean_data, {
                'tipo_animal': 'tipo_id',
                'status': 'status_id',
                'pasto': 'pasto_id',
                'raca': 'raca_id',
                'origem': 'origem_id'
            })
            self.db_manager.insert('animais', clean_data)

        elif tipo == "Talhões":
            if 'data_plantio' in clean_data:
                clean_data['data_plantio'] = self.parse_date(clean_data['data_plantio'])

            # Mapear variedade
            clean_data = self.map_foreign_keys(clean_data, {
                'variedade': 'variedade_id'
            })
            self.db_manager.insert('talhoes', clean_data)

        elif tipo == "Despesas":
            # Converter datas
            if 'data_gasto' in clean_data:
                clean_data['data_gasto'] = self.parse_date(clean_data['data_gasto'])
            if 'data_vencimento' in clean_data:
                clean_data['data_vencimento'] = self.parse_date(clean_data['data_vencimento'])
            if 'data_pagamento' in clean_data:
                clean_data['data_pagamento'] = self.parse_date(clean_data['data_pagamento'])

            # Mapear foreign keys
            clean_data = self.map_foreign_keys(clean_data, {
                'tipo_despesa': 'tipo_despesa_id',
                'fornecedor': 'fornecedor_id',
                'conta_bancaria': 'conta_bancaria_id'
            })

            # Calcular valor_final se não fornecido
            if 'valor_final' not in clean_data:
                qtd = float(clean_data.get('quantidade', 1))
                val_unit = float(clean_data.get('valor_unitario', clean_data.get('valor', 0)))
                desc = float(clean_data.get('desconto', 0))
                clean_data['valor_final'] = (qtd * val_unit) - desc

            self.db_manager.insert('despesas', clean_data)

        elif tipo == "Receitas":
            # Converter datas
            if 'data_venda' in clean_data:
                clean_data['data_venda'] = self.parse_date(clean_data['data_venda'])
            if 'data_vencimento' in clean_data:
                clean_data['data_vencimento'] = self.parse_date(clean_data['data_vencimento'])
            if 'data_pagamento' in clean_data:
                clean_data['data_pagamento'] = self.parse_date(clean_data['data_pagamento'])

            # Mapear foreign keys
            clean_data = self.map_foreign_keys(clean_data, {
                'tipo_receita': 'tipo_receita_id',
                'cliente': 'cliente_id',
                'conta_bancaria': 'conta_bancaria_id'
            })

            # Calcular valor_total se não fornecido
            if 'valor_total' not in clean_data:
                qtd = float(clean_data.get('quantidade', 1))
                val_unit = float(clean_data.get('valor_unitario', clean_data.get('valor', 0)))
                desc = float(clean_data.get('desconto', 0))
                clean_data['valor_total'] = (qtd * val_unit) - desc
                clean_data['valor_final'] = clean_data['valor_total']

            self.db_manager.insert('receitas', clean_data)

        elif tipo == "Tratos Culturais":
            if 'data_execucao' in clean_data:
                clean_data['data_execucao'] = self.parse_date(clean_data['data_execucao'])
            if 'proxima_aplicacao' in clean_data:
                clean_data['proxima_aplicacao'] = self.parse_date(clean_data['proxima_aplicacao'])

            clean_data = self.map_foreign_keys(clean_data, {
                'talhao': 'talhao_id',
                'responsavel': 'responsavel_id'
            })
            self.db_manager.insert('tratos_culturais', clean_data)

        elif tipo == "Colheitas":
            if 'data_colheita' in clean_data:
                clean_data['data_colheita'] = self.parse_date(clean_data['data_colheita'])

            clean_data = self.map_foreign_keys(clean_data, {
                'talhao': 'talhao_id',
                'responsavel': 'responsavel_id'
            })
            self.db_manager.insert('colheitas_banana', clean_data)

    def map_foreign_keys(self, data, mapping):
        """Mapeia nomes para IDs de foreign keys"""
        result = data.copy()

        for name_field, id_field in mapping.items():
            if name_field in result:
                name_value = str(result[name_field]).strip()

                # Determinar tabela de lookup baseado no campo
                lookup_table = None
                search_column = 'nome'  # Padrão

                if 'tipo_despesa' in id_field:
                    lookup_table = 'tipo_despesa'
                elif 'tipo_receita' in id_field:
                    lookup_table = 'tipo_receita'
                elif 'tipo_' in id_field:
                    lookup_table = 'tipo_animal'
                elif 'status' in id_field:
                    lookup_table = 'status_animal'
                elif 'pasto' in id_field:
                    lookup_table = 'pastos'
                elif 'raca' in id_field:
                    lookup_table = 'raca'
                elif 'origem' in id_field:
                    lookup_table = 'origem'
                elif 'fornecedor' in id_field:
                    lookup_table = 'fornecedores'
                elif 'cliente' in id_field:
                    lookup_table = 'clientes'
                elif 'conta_bancaria' in id_field:
                    lookup_table = 'contas_bancarias'
                    search_column = 'nome_conta'
                elif 'variedade' in id_field:
                    lookup_table = 'variedades_banana'
                elif 'talhao' in id_field:
                    lookup_table = 'talhoes'
                    search_column = 'codigo'
                elif 'responsavel' in id_field:
                    lookup_table = 'funcionarios'

                if lookup_table:
                    # Buscar ID pelo nome ou código
                    try:
                        # Primeiro tenta buscar apenas por nome/codigo
                        lookup_result = self.db_manager.select(
                            lookup_table,
                            'id',
                            f'{search_column} = ?',
                            [name_value]
                        )

                        if lookup_result:
                            result[id_field] = lookup_result[0]['id']
                            del result[name_field]
                        else:
                            # Se não encontrou, tentar criar um novo registro básico
                            if lookup_table in ['tipo_despesa', 'tipo_receita', 'pastos', 'raca', 'origem']:
                                try:
                                    # Verifica se já existe antes de inserir
                                    existing = self.db_manager.select(
                                        lookup_table,
                                        'id',
                                        'nome = ?',
                                        [name_value]
                                    )

                                    if existing:
                                        result[id_field] = existing[0]['id']
                                    else:
                                        new_id = self.db_manager.insert(lookup_table, {'nome': name_value})
                                        if new_id:
                                            result[id_field] = new_id

                                    del result[name_field]
                                except Exception as e:
                                    print(f"Aviso: Não foi possível mapear {name_field}={name_value}: {e}")
                                    # Remove o campo original para evitar erro de sintaxe
                                    del result[name_field]
                            else:
                                # Para outras tabelas, se não existe, remove o campo
                                print(f"Aviso: {name_field}={name_value} não encontrado em {lookup_table}")
                                del result[name_field]

                    except Exception as e:
                        print(f"Erro ao mapear {name_field}: {e}")
                        # Remove o campo problemático
                        if name_field in result:
                            del result[name_field]

        return result

    def parse_date(self, date_value):
        """Converte data do Excel para formato SQLite"""
        if isinstance(date_value, datetime):
            return date_value.strftime('%Y-%m-%d')
        elif isinstance(date_value, str):
            # Tentar diversos formatos
            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                try:
                    return datetime.strptime(date_value, fmt).strftime('%Y-%m-%d')
                except:
                    continue
        return date_value

    def download_template(self):
        """Baixa modelo Excel com seleção de colunas"""
        tipo = self.tipo_combo.get()
        if not tipo:
            messagebox.showwarning("Aviso", "Selecione o tipo de dados primeiro!")
            return

        self.update_mapping_template()

        # Criar janela de seleção de colunas
        dialog = tk.Toplevel(self.parent)
        dialog.title(f"Selecionar Colunas para {tipo}")
        dialog.geometry("500x600")
        dialog.transient(self.parent)
        dialog.grab_set()

        # Frame principal com scroll
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=BOTH, expand=YES)

        ttk.Label(main_frame, text="Selecione as colunas que deseja incluir no modelo:",
                 font=("Helvetica", 12, "bold")).pack(pady=(0, 10))

        ttk.Label(main_frame, text="⚠ Campos marcados com * são obrigatórios",
                 foreground="red").pack(pady=(0, 10))

        # Canvas para scroll
        canvas = tk.Canvas(main_frame, height=400)
        scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Checkboxes para cada coluna
        column_vars = {}
        for field_name, is_required, example in self.current_template:
            var = tk.BooleanVar(value=True if is_required else False)
            column_vars[field_name] = (var, is_required, example)

            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=X, padx=10, pady=5)

            check = ttk.Checkbutton(
                frame,
                text=f"{field_name}{'*' if is_required else ''}",
                variable=var,
                state="disabled" if is_required else "normal"
            )
            check.pack(side=LEFT)

            ttk.Label(frame, text=f"  (ex: {example})",
                     foreground="gray").pack(side=LEFT)

        canvas.pack(side=LEFT, fill=BOTH, expand=YES)
        scrollbar.pack(side=RIGHT, fill=Y)

        # Botões
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=X, pady=20)

        def select_all():
            for var, _, _ in column_vars.values():
                var.set(True)

        def deselect_optional():
            for field_name, (var, is_required, _) in column_vars.items():
                if not is_required:
                    var.set(False)

        ttk.Button(btn_frame, text="Selecionar Tudo", command=select_all).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="Apenas Obrigatórios", command=deselect_optional).pack(side=LEFT, padx=5)

        def generate_template():
            # Coletar colunas selecionadas
            selected_columns = [
                (field_name, example)
                for field_name, (var, _, example) in column_vars.items()
                if var.get()
            ]

            if not selected_columns:
                messagebox.showwarning("Aviso", "Selecione pelo menos uma coluna!")
                return

            filename = filedialog.asksaveasfilename(
                title="Salvar modelo",
                defaultextension=".xlsx",
                initialfile=f"modelo_{tipo.lower().replace(' ', '_')}.xlsx",
                filetypes=[("Excel", "*.xlsx")]
            )

            if filename:
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = tipo[:31]  # Excel tem limite de 31 caracteres no nome da sheet

                    # Estilizar cabeçalho
                    from openpyxl.styles import Font, PatternFill, Alignment

                    # Adicionar cabeçalhos
                    for idx, (col_name, example) in enumerate(selected_columns, 1):
                        cell = ws.cell(1, idx, col_name)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")

                        # Adicionar linha de exemplo
                        ws.cell(2, idx, example)

                        # Ajustar largura da coluna
                        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max(len(col_name), len(str(example))) + 5

                    wb.save(filename)
                    dialog.destroy()
                    messagebox.showinfo("Sucesso",
                                       f"Modelo salvo com {len(selected_columns)} colunas!\n\n{filename}")

                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao criar modelo: {str(e)}")

        ttk.Button(main_frame, text="Gerar Modelo Excel", command=generate_template,
                  width=25).pack(pady=10)
        ttk.Button(main_frame, text="Cancelar", command=dialog.destroy,
                  width=25).pack()
